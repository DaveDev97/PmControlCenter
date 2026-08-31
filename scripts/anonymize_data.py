"""
Generate anonymized Excel files for demo/testing purposes.

Reads from ../backend/app/seed.py data structure and creates:
- contracts_financials.xlsx
- allocations.xlsx
- opportunities.xlsx

All data is anonymized (fake names, companies, amounts scaled randomly).
"""
import random
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


# Fake names pools
FIRST_NAMES = [
    "Manager", "Senior Consultant", "Consultant", "Analyst", "Lead",
    "Director", "Specialist", "Coordinator", "Architect", "Engineer"
]

SUFFIXES = ["A", "B", "C", "D", "E", "F", "1", "2", "3", "4"]

CLIENTS = [
    "Alpha Corp", "Beta Industries", "Gamma Ltd", "Delta Group",
    "Epsilon Holdings", "Zeta Partners", "Theta Solutions", "Iota Ventures"
]

LEGAL_ENTITIES = [
    "Alpha Corp S.p.A.", "Beta Industries Ltd.", "Gamma Group Inc.",
    "Delta Holdings S.p.A.", "Epsilon Partners S.r.l."
]

SERVICE_TYPES = [
    "Security Services", "Cloud Security", "SOC Operations",
    "Vulnerability Management", "IAM Services", "Data Protection",
    "Threat Intelligence", "Security Consulting"
]

# Output folder
OUTPUT_DIR = Path(__file__).parent.parent / "sample_data"
OUTPUT_DIR.mkdir(exist_ok=True)


def random_scale(value: float, min_factor=0.7, max_factor=1.3) -> float:
    """Scale amount randomly to anonymize."""
    return round(value * random.uniform(min_factor, max_factor), 2)


def generate_resources(count=20) -> pd.DataFrame:
    """Generate fake resources."""
    resources = []

    roles_pool = [
        ("Manager", 800.0),
        ("Senior Consultant", 450.0),
        ("Consultant", 350.0),
        ("Analyst", 280.0),
    ]

    for i in range(count):
        role_idx = i % len(roles_pool)
        role, base_rate = roles_pool[role_idx]

        name = f"{role} {SUFFIXES[i % len(SUFFIXES)]}"
        email = f"{name.lower().replace(' ', '.')}@company.com"

        resources.append({
            "name": name,
            "email": email,
            "role": role,
            "daily_rate": round(base_rate * random.uniform(0.9, 1.1), 2),
            "loaded_cost_hourly": round(base_rate / 8 * random.uniform(0.8, 1.0), 2),
            "chargeability": round(random.uniform(0.70, 0.90), 2),
            "status": "active",
            "hire_date": date(2024 + random.randint(0, 2), random.randint(1, 12), 1)
        })

    return pd.DataFrame(resources)


def generate_contracts(count=5) -> pd.DataFrame:
    """Generate fake contracts."""
    contracts = []

    for i in range(count):
        client = random.choice(CLIENTS)
        service = random.choice(SERVICE_TYPES)

        contracts.append({
            "contract_id": f"PROJ-{str(i+1).zfill(3)}",
            "client_name": client,
            "contract_name": f"{client.split()[0]} - {service}",
            "service_group": random.choice(["IMS", "AMS", "Security", "Consulting"]),
            "wbs_l1": f"WBS{str(i+1).zfill(3)}",
            "wbs_l2": f"SUB{str(i+1).zfill(3)}",
            "description": f"Managed {service.lower()} for {client}",
            "contract_type": random.choice(["T&M", "Fixed Price", "Capped T&M"]),
            "fiscal_year": random.choice(["2025", "2026"]),
            "start_date": date(2026, 1, 1),
            "end_date": date(2026, 12, 31),
            "initial_budget": random_scale(800000.0, 0.8, 1.5),
            "status": "active"
        })

    return pd.DataFrame(contracts)


def generate_financials(contracts: pd.DataFrame) -> pd.DataFrame:
    """Generate monthly financials for each contract."""
    financials = []

    months = pd.date_range("2026-01-01", "2026-12-01", freq="MS")

    for _, contract in contracts.iterrows():
        budget = contract["initial_budget"]
        monthly_target = budget / 12

        for month in months:
            # Actual months (Jan-Aug)
            if month.month <= 8:
                revenues_actual = random_scale(monthly_target, 0.7, 1.2)
                costs_actual = revenues_actual * random.uniform(0.55, 0.75)  # 25-45% margin

                financials.append({
                    "contract_id": contract["contract_id"],
                    "month": month.date(),
                    "fiscal_quarter": f"Q{((month.month - 1) // 3) + 1}",
                    "is_actual": True,
                    "billings_actual": revenues_actual,
                    "revenues_actual": revenues_actual,
                    "payroll_costs_actual": costs_actual * 0.85,
                    "non_payroll_costs_actual": costs_actual * 0.10,
                    "capital_charges_actual": costs_actual * 0.05,
                    "billings_forecast": 0.0,
                    "revenues_forecast": 0.0,
                    "payroll_costs_forecast": 0.0,
                    "non_payroll_costs_forecast": 0.0,
                    "capital_charges_forecast": 0.0
                })
            # Forecast months (Sep-Dec)
            else:
                revenues_forecast = random_scale(monthly_target, 0.8, 1.1)
                costs_forecast = revenues_forecast * random.uniform(0.60, 0.70)

                financials.append({
                    "contract_id": contract["contract_id"],
                    "month": month.date(),
                    "fiscal_quarter": f"Q{((month.month - 1) // 3) + 1}",
                    "is_actual": False,
                    "billings_actual": 0.0,
                    "revenues_actual": 0.0,
                    "payroll_costs_actual": 0.0,
                    "non_payroll_costs_actual": 0.0,
                    "capital_charges_actual": 0.0,
                    "billings_forecast": revenues_forecast,
                    "revenues_forecast": revenues_forecast,
                    "payroll_costs_forecast": costs_forecast * 0.85,
                    "non_payroll_costs_forecast": costs_forecast * 0.10,
                    "capital_charges_forecast": costs_forecast * 0.05
                })

    return pd.DataFrame(financials)


def generate_allocations(resources: pd.DataFrame, contracts: pd.DataFrame) -> pd.DataFrame:
    """Generate resource allocations."""
    allocations = []

    # Assegna ogni risorsa a 1-3 contratti
    for _, resource in resources.iterrows():
        num_contracts = random.randint(1, 3)
        assigned_contracts = random.sample(list(contracts["contract_id"]), min(num_contracts, len(contracts)))

        for contract_id in assigned_contracts:
            utilization = round(random.uniform(0.4, 1.0), 2)
            days_per_month = round(utilization * 22, 1)  # 22 giorni lavorativi/mese

            allocations.append({
                "resource_name": resource["name"],
                "resource_email": resource["email"],
                "contract_id": contract_id,
                "utilization": utilization,
                "days_per_month": days_per_month,
                "start_date": date(2026, 1, 1),
                "end_date": date(2026, 12, 31)
            })

    return pd.DataFrame(allocations)


def generate_opportunities(contracts: pd.DataFrame) -> pd.DataFrame:
    """Generate fake opportunities."""
    opportunities = []

    stages = ["Lead", "Qualified", "Proposal", "CloseWon"]

    for i in range(10):
        contract = random.choice(contracts["contract_id"].tolist())
        client = contracts[contracts["contract_id"] == contract]["client_name"].iloc[0]
        service = random.choice(SERVICE_TYPES)
        stage = random.choice(stages)

        opportunities.append({
            "opp_id_mms": f"OPP-{str(i+1).zfill(4)}",
            "contract_id": contract if stage == "CloseWon" else None,
            "name": f"{client.split()[0]} {service}",
            "description": f"New opportunity for {service.lower()} with {client}",
            "legal_entity": random.choice(LEGAL_ENTITIES),
            "fiscal_year": random.choice(["2026", "2027"]),
            "close_date": date(2026, random.randint(9, 12), random.randint(1, 28)),
            "quarter": random.choice(["Q3", "Q4"]),
            "pds_status": random.choice(["Da inviare", "Inviata", "Approvata CVS"]),
            "acn_tool_status": random.choice(["Todo", "In progress", "Done"]),
            "mms_status": stage,
            "stage": stage,
            "estimated_value": random_scale(150000.0, 0.5, 2.0),
            "probability": 1.0 if stage == "CloseWon" else (0.6 if stage == "Proposal" else 0.3),
            "notes": None
        })

    return pd.DataFrame(opportunities)


def style_excel_sheet(ws, df: pd.DataFrame, title: str):
    """Apply styling to Excel sheet."""
    # Header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 50)

    # Freeze header
    ws.freeze_panes = "A2"


def create_contracts_financials_excel():
    """Create contracts_financials.xlsx with 3 sheets."""
    print("📊 Generating contracts_financials.xlsx...")

    # Generate data
    resources_df = generate_resources(20)
    contracts_df = generate_contracts(5)
    financials_df = generate_financials(contracts_df)

    # Create Excel
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Contracts
    ws_contracts = wb.create_sheet("Contracts")
    for row in dataframe_to_rows(contracts_df, index=False, header=True):
        ws_contracts.append(row)
    style_excel_sheet(ws_contracts, contracts_df, "Contracts")

    # Sheet 2: Financials
    ws_financials = wb.create_sheet("Financials")
    for row in dataframe_to_rows(financials_df, index=False, header=True):
        ws_financials.append(row)
    style_excel_sheet(ws_financials, financials_df, "Financials")

    # Sheet 3: Resources
    ws_resources = wb.create_sheet("Resources")
    for row in dataframe_to_rows(resources_df, index=False, header=True):
        ws_resources.append(row)
    style_excel_sheet(ws_resources, resources_df, "Resources")

    # Save
    output_path = OUTPUT_DIR / "contracts_financials.xlsx"
    wb.save(output_path)
    print(f"✅ Created: {output_path}")
    print(f"   - {len(contracts_df)} contracts")
    print(f"   - {len(financials_df)} financial records")
    print(f"   - {len(resources_df)} resources")

    return contracts_df, resources_df


def create_allocations_excel(resources_df: pd.DataFrame, contracts_df: pd.DataFrame):
    """Create allocations.xlsx."""
    print("\n📊 Generating allocations.xlsx...")

    allocations_df = generate_allocations(resources_df, contracts_df)

    wb = Workbook()
    ws = wb.active
    ws.title = "Allocations"

    for row in dataframe_to_rows(allocations_df, index=False, header=True):
        ws.append(row)

    style_excel_sheet(ws, allocations_df, "Allocations")

    output_path = OUTPUT_DIR / "allocations.xlsx"
    wb.save(output_path)
    print(f"✅ Created: {output_path}")
    print(f"   - {len(allocations_df)} allocations")


def create_opportunities_excel(contracts_df: pd.DataFrame):
    """Create opportunities.xlsx."""
    print("\n📊 Generating opportunities.xlsx...")

    opportunities_df = generate_opportunities(contracts_df)

    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"

    for row in dataframe_to_rows(opportunities_df, index=False, header=True):
        ws.append(row)

    style_excel_sheet(ws, opportunities_df, "Opportunities")

    output_path = OUTPUT_DIR / "opportunities.xlsx"
    wb.save(output_path)
    print(f"✅ Created: {output_path}")
    print(f"   - {len(opportunities_df)} opportunities")


def create_readme():
    """Create README.md in sample_data folder."""
    readme_content = """# Sample Data - Excel Schema

Questi file Excel contengono dati **anonimizzati** per testing/demo dell'applicazione.

## 📁 File Struttura

### 1. contracts_financials.xlsx

#### Sheet: Contracts
| Colonna | Tipo | Descrizione | Esempio |
|---------|------|-------------|---------|
| contract_id | String | ID univoco contratto | PROJ-001 |
| client_name | String | Nome cliente | Alpha Corp |
| contract_name | String | Nome contratto | Alpha - Security Services |
| service_group | String | Gruppo servizio | IMS, AMS, Security |
| wbs_l1 | String | WBS livello 1 | WBS001 |
| wbs_l2 | String | WBS livello 2 | SUB001 |
| description | String | Descrizione | Managed security services... |
| contract_type | String | Tipo | T&M, Fixed Price, Capped T&M |
| fiscal_year | String | Anno fiscale | 2026 |
| start_date | Date | Data inizio | 2026-01-01 |
| end_date | Date | Data fine | 2026-12-31 |
| initial_budget | Float | Budget iniziale (EUR) | 950000.00 |
| status | String | Stato | active, closed |

#### Sheet: Financials
| Colonna | Tipo | Descrizione | Esempio |
|---------|------|-------------|---------|
| contract_id | String | Riferimento contratto | PROJ-001 |
| month | Date | Mese (primo giorno) | 2026-01-01 |
| fiscal_quarter | String | Quarter fiscale | Q1, Q2, Q3, Q4 |
| is_actual | Boolean | Actual vs Forecast | TRUE, FALSE |
| billings_actual | Float | Fatturato actual | 78000.00 |
| revenues_actual | Float | Ricavi actual | 78000.00 |
| payroll_costs_actual | Float | Costi personale actual | 52000.00 |
| non_payroll_costs_actual | Float | Costi non personale actual | 3000.00 |
| capital_charges_actual | Float | Ammortamenti actual | 2000.00 |
| *_forecast | Float | Stesse colonne per forecast | ... |

#### Sheet: Resources
| Colonna | Tipo | Descrizione | Esempio |
|---------|------|-------------|---------|
| name | String | Nome risorsa | Manager A |
| email | String | Email | manager.a@company.com |
| role | String | Ruolo | Manager, Senior Consultant |
| daily_rate | Float | Tariffa giornaliera | 800.00 |
| loaded_cost_hourly | Float | Costo orario caricato | 95.00 |
| chargeability | Float | Chargeability target | 0.85 |
| status | String | Stato | active, inactive |
| hire_date | Date | Data assunzione | 2024-01-01 |

---

### 2. allocations.xlsx

#### Sheet: Allocations
| Colonna | Tipo | Descrizione | Esempio |
|---------|------|-------------|---------|
| resource_name | String | Nome risorsa | Manager A |
| resource_email | String | Email risorsa | manager.a@company.com |
| contract_id | String | ID contratto | PROJ-001 |
| utilization | Float | Utilizzo (0-1) | 0.85 |
| days_per_month | Float | Giorni/mese | 18.0 |
| start_date | Date | Data inizio allocazione | 2026-01-01 |
| end_date | Date | Data fine allocazione | 2026-12-31 |

---

### 3. opportunities.xlsx

#### Sheet: Opportunities
| Colonna | Tipo | Descrizione | Esempio |
|---------|------|-------------|---------|
| opp_id_mms | String | ID opportunità MMS | OPP-0001 |
| contract_id | String | Contratto collegato (se CloseWon) | PROJ-001 |
| name | String | Nome opportunità | Alpha Security Extension |
| description | String | Descrizione | Extension servizi SOC |
| legal_entity | String | Legal entity cliente | Alpha Corp S.p.A. |
| fiscal_year | String | Anno fiscale | 2026 |
| close_date | Date | Data chiusura prevista | 2026-10-31 |
| quarter | String | Quarter | Q1, Q2, Q3, Q4 |
| pds_status | String | Status PDS | Da inviare, Inviata, Approvata CVS |
| acn_tool_status | String | Status ACN Tool | Todo, In progress, Done |
| mms_status | String | Status MMS | Lead, Qualified, Proposal, CloseWon |
| stage | String | Stage opportunità | Lead, Qualified, Proposal, CloseWon |
| estimated_value | Float | Valore stimato (EUR) | 120000.00 |
| probability | Float | Probabilità (0-1) | 0.6 |
| notes | String | Note (opzionale) | Follow-up Q3 |

---

## 🔐 Anonimizzazione

Questi dati sono **completamente fittizi**:

- ✅ Nomi persone → "Manager A", "Senior Consultant 1"
- ✅ Email → "role.suffix@company.com"
- ✅ Clienti → "Alpha Corp", "Beta Industries"
- ✅ Contratti → "PROJ-001", "PROJ-002"
- ✅ Importi → scalati random (0.7x - 1.3x)

**Nessun dato reale incluso.**

---

## 🎯 Utilizzo

1. **Setup Wizard**: Punta alla cartella `sample_data/`
2. **App carica** i 3 Excel automaticamente
3. **Dashboard popolate** con dati demo
4. **Modifica DD/Overlay** → salvato in `pm_overlay.json`

---

**Generato da:** `scripts/anonymize_data.py`
**Data:** {datetime.now().strftime('%Y-%m-%d')}
**Versione:** 1.0.0
"""

    readme_path = OUTPUT_DIR / "README.md"
    readme_path.write_text(readme_content)
    print(f"\n✅ Created: {readme_path}")


def main():
    """Generate all anonymized Excel files."""
    print("="*60)
    print("  PM Control Center - Data Anonymization")
    print("="*60)

    random.seed(42)  # Reproducible results

    # Generate files
    contracts_df, resources_df = create_contracts_financials_excel()
    create_allocations_excel(resources_df, contracts_df)
    create_opportunities_excel(contracts_df)
    create_readme()

    print("\n" + "="*60)
    print("✅ All sample data files generated successfully!")
    print(f"📁 Output folder: {OUTPUT_DIR}")
    print("="*60)


if __name__ == "__main__":
    main()
