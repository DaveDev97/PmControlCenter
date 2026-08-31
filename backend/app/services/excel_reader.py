"""Load PM Control Center data from the real security-financials workbook.

The production data source is a single Excel workbook with the same structure as
``BNL_Security_Financials_v02.xlsx`` (16 sheets, irregular layouts). The shipped
sample ``security_financials.xlsx`` is a structure-preserving anonymized clone of
it (see ``scripts/anonymize_data.py``), so this reader works identically on both.

Only the sheets the application needs are parsed:

* ``Contracts``          -> Contracts + monthly Financials (per-contract blocks)
* ``Costi vs Forecast``  -> Resources (loaded cost, chargeability) + Allocations
* ``Opp. FY25/26/27``    -> Opportunities

Sheets are located by tolerant name matching so anonymized titles (which may
rename embedded client tokens) still resolve.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Allocation,
    Client,
    Contract,
    Financial,
    Opportunity,
    Resource,
)

DATA_FILENAME = "security_financials.xlsx"
REQUIRED_FILES = (DATA_FILENAME,)
OPTIONAL_FILES: tuple[str, ...] = ()

_CONTRACT_HEADER_RE = re.compile(r"^\s*(\d{6,})\s*-\s*(.+?)\s*$")
_METRIC_MAP = {
    "billing": "billings",
    "billings": "billings",
    "revenue": "revenues",
    "revenues": "revenues",
    "payroll": "payroll_costs",
    "payroll costs": "payroll_costs",
    "non payroll": "non_payroll_costs",
    "non payroll costs": "non_payroll_costs",
    "capital charge": "capital_charges",
    "capital charges": "capital_charges",
}
_BACKFILL_MONTH = date(2026, 1, 1)  # "Previous" cumulative parked here for YTD realism


# --------------------------------------------------------------------------- #
# coercion helpers
# --------------------------------------------------------------------------- #
def _is_blank(v) -> bool:
    return v is None or v == "" or (isinstance(v, float) and v != v)


def _s(v) -> str | None:
    return None if _is_blank(v) else str(v).strip()


def _f(v, default: float = 0.0) -> float:
    if _is_blank(v):
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _d(v) -> date | None:
    if _is_blank(v):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


# --------------------------------------------------------------------------- #
class ExcelDataLoader:
    """Loads data from the security-financials workbook into the database."""

    @staticmethod
    def resolve_workbook(path) -> Path | None:
        """Resolve the data source to an actual .xlsx file.

        Accepts either the Excel file directly (any name — no constraint) or, for
        backward compatibility, a folder (uses security_financials.xlsx, else the
        first .xlsx found).
        """
        p = Path(path)
        if p.is_file():
            return p
        if p.is_dir():
            cand = p / DATA_FILENAME
            if cand.exists():
                return cand
            xlsx = sorted(p.glob("*.xlsx"))
            return xlsx[0] if xlsx else None
        return None

    def validate_folder(self, data_folder: Path) -> dict:
        """Validate the selected data source by CONTENT (not by file name)."""
        wb_path = self.resolve_workbook(data_folder)
        if wb_path is None:
            return {"valid": False, "found": [], "missing": ["file Excel"], "optional": [],
                    "error": f"Nessun file Excel trovato: {data_folder}"}
        try:
            wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
            titles = {t.lower() for t in wb.sheetnames}
            wb.close()
        except Exception as exc:  # noqa: BLE001 - report any open error to the user
            return {"valid": False, "found": [wb_path.name], "missing": [], "optional": [],
                    "error": f"Impossibile aprire il file: {exc}"}
        looks_valid = ("contracts" in titles or any(t.startswith("opp.") for t in titles)
                       or any("forecast" in t for t in titles))
        if not looks_valid:
            return {"valid": False, "found": [wb_path.name], "missing": [], "optional": [],
                    "error": "Il file non sembra un workbook PM Control Center "
                             "(mancano i fogli Contracts / Opp. / Forecast)."}
        return {"valid": True, "found": [wb_path.name], "missing": [], "optional": [], "path": str(wb_path)}

    async def load_all(self, data_folder: Path, session: AsyncSession) -> dict:
        wb_path = self.resolve_workbook(data_folder)
        if wb_path is None:
            raise FileNotFoundError(f"No Excel file at {data_folder}")
        wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
        counts = {k: 0 for k in
                  ("clients", "contracts", "financials", "resources", "allocations", "opportunities")}

        contract_ids = await self._load_contracts(wb, session, counts)
        await self._load_resources_and_allocations(wb, session, counts, contract_ids)
        await self._load_opportunities(wb, session, counts, contract_ids)

        await session.commit()
        return counts

    # ------------------------------------------------------------------ #
    def _sheet(self, wb, *predicates):
        """Return a worksheet by title, trying predicates in priority order."""
        for pred in predicates:
            for title in wb.sheetnames:
                if pred(title.lower()):
                    return wb[title]
        return None

    async def _load_contracts(self, wb, session, counts) -> set[str]:
        ws = self._sheet(wb, lambda t: t == "contracts")
        if ws is None:
            return set()
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            return set()

        # Month columns from the first two header rows.
        header, subhdr = rows[0], (rows[1] if len(rows) > 1 else [])
        month_cols: list[tuple[int, date, bool]] = []
        for idx, val in enumerate(header):
            d = val.date() if isinstance(val, datetime) else None
            if d is None:
                continue
            sub = subhdr[idx] if idx < len(subhdr) else None
            is_actual = str(sub).strip().lower() == "actual" if sub else False
            month_cols.append((idx, d.replace(day=1), is_actual))
        prev_col = next((i for i, v in enumerate(header) if _s(v) == "Previous"), None)

        clients: dict[str, int] = {}
        contract_ids: set[str] = set()
        i, n = 0, len(rows)
        while i < n:
            head = _s(rows[i][0]) if rows[i] else None
            m = _CONTRACT_HEADER_RE.match(head) if head else None
            if not m:
                i += 1
                continue
            contract_id, name = m.group(1), m.group(2)
            # Optional WBS on the following row (single token in col A).
            wbs = None
            if i + 1 < n:
                nxt = _s(rows[i + 1][0])
                if nxt and " " not in nxt and not _CONTRACT_HEADER_RE.match(nxt):
                    wbs = nxt

            # Collect metric rows until the next contract header (or 12 rows).
            metrics: dict[str, list] = {}
            j = i + 1
            while j < n and j < i + 13:
                label = _s(rows[j][0])
                if label and _CONTRACT_HEADER_RE.match(label):
                    break
                key = _METRIC_MAP.get((label or "").lower())
                if key:
                    metrics[key] = rows[j]
                j += 1

            if metrics:
                client_name = self._client_name(name, len(clients))
                if client_name not in clients:
                    c = Client(name=client_name, industry="Financial Services")
                    session.add(c)
                    await session.flush()
                    clients[client_name] = c.id
                    counts["clients"] += 1

                session.add(Contract(
                    id=contract_id, client_id=clients[client_name], name=name,
                    wbs_l1=wbs, contract_type="T&M", fiscal_year="FY26",
                    start_date=date(2025, 9, 1), end_date=date(2026, 8, 31), status="active",
                ))
                contract_ids.add(contract_id)
                counts["contracts"] += 1

                # Backfill "Previous" cumulative as one actual month.
                if prev_col is not None:
                    fin = self._financial(contract_id, _BACKFILL_MONTH, True, metrics, prev_col)
                    if fin:
                        session.add(fin)
                        counts["financials"] += 1
                for col, d, is_actual in month_cols:
                    fin = self._financial(contract_id, d, is_actual, metrics, col)
                    if fin:
                        session.add(fin)
                        counts["financials"] += 1
            i = j

        await session.flush()
        return contract_ids

    @staticmethod
    def _client_name(contract_name: str, index: int) -> str:
        first = (contract_name or "").split()[0] if contract_name else ""
        if first[:1].isupper() and first.isalpha() and len(first) >= 4:
            return first
        return f"Cliente {index + 1}"

    @staticmethod
    def _financial(contract_id, month, is_actual, metrics, col) -> Financial | None:
        vals = {}
        any_val = False
        for key, row in metrics.items():
            v = _f(row[col]) if col < len(row) else 0.0
            if v:
                any_val = True
            vals[key] = v
        if not any_val:
            return None
        suffix = "actual" if is_actual else "forecast"
        return Financial(
            contract_id=contract_id, month=month, is_actual=is_actual,
            fiscal_quarter=f"Q{((month.month - 1) // 3) + 1}",
            **{f"{k}_{suffix}": vals.get(k, 0.0) for k in
               ("billings", "revenues", "payroll_costs", "non_payroll_costs", "capital_charges")},
        )

    async def _load_resources_and_allocations(self, wb, session, counts, contract_ids):
        ws = self._sheet(wb, lambda t: "costi vs forecast" in t, lambda t: "forecast 26" in t)
        if ws is None:
            return
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        # Locate the header row that starts the resource table ("Resource" in a col).
        hdr_idx = res_col = lc_col = charg_col = None
        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                if _s(val) == "Resource":
                    hdr_idx, res_col = ri, ci
                    for cj in range(ci + 1, min(ci + 5, len(row))):
                        lbl = (_s(row[cj]) or "").lower()
                        if lbl == "lc":
                            lc_col = cj
                        elif lbl.startswith("%charg") or lbl == "charg":
                            charg_col = cj
                    break
            if hdr_idx is not None:
                break
        if hdr_idx is None:
            return

        first_contract = next(iter(contract_ids), None)
        seen: set[str] = set()
        for row in rows[hdr_idx + 1:]:
            name = _s(row[res_col]) if res_col < len(row) else None
            if not name or "." not in name:
                continue
            if name in seen:
                continue
            seen.add(name)
            lc = _f(row[lc_col]) if lc_col is not None and lc_col < len(row) else 0.0
            charg = _f(row[charg_col], 0.80) if charg_col is not None and charg_col < len(row) else 0.80
            res = Resource(
                name=name, email=f"{name}@example.com",
                daily_rate=round(lc * 8, 2), loaded_cost_hourly=lc or None,
                chargeability=charg, status="active",
            )
            session.add(res)
            await session.flush()
            counts["resources"] += 1
            if first_contract:
                session.add(Allocation(
                    resource_id=res.id, contract_id=first_contract,
                    utilization=charg, days_per_month=round(charg * 20),
                    start_date=date(2026, 1, 1), end_date=date(2026, 8, 31),
                ))
                counts["allocations"] += 1

    async def _load_opportunities(self, wb, session, counts, contract_ids):
        for title in wb.sheetnames:
            if not title.lower().startswith("opp."):
                continue
            ws = wb[title]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            # Header row = the one containing "Contract".
            hdr_idx = None
            for ri, row in enumerate(rows[:5]):
                if any(_s(v) == "Contract" for v in row):
                    hdr_idx = ri
                    break
            if hdr_idx is None:
                continue
            cols = {(_s(v) or "").lower(): ci for ci, v in enumerate(rows[hdr_idx])}

            def g(row, *names):
                for nm in names:
                    ci = cols.get(nm)
                    if ci is not None and ci < len(row):
                        return row[ci]
                return None

            for row in rows[hdr_idx + 1:]:
                name = _s(g(row, "project"))
                opp_id = _s(g(row, "opp id mms"))
                if not name and not opp_id:
                    continue
                cid = _s(g(row, "contract"))
                if cid not in contract_ids:
                    cid = None
                mms = _s(g(row, "mms status")) or "Lead"
                stage = self._stage(mms)
                session.add(Opportunity(
                    opp_id_mms=opp_id, contract_id=cid, name=name or (opp_id or "Opportunity"),
                    legal_entity=None, fiscal_year=self._fy(title),
                    close_date=_d(g(row, "close date")),
                    quarter=_s(g(row, "close date quarter")),
                    pds_status=_s(g(row, "pds status")),
                    acn_tool_status=_s(g(row, "stato acn tool")),
                    mms_status=mms, stage=stage,
                    oda_id=_s(g(row, "oda id")),
                    ccp_number=_s(g(row, "alphabank contract (ccp)", "bnl contract (ccp)")),
                    referente=_s(g(row, "ref name")),
                    mmr_code=_s(g(row, "mmr code")),
                    estimated_value=_f(g(row, "revenues")),
                    total_invoiced=_f(g(row, "billed")),
                    total_to_invoice=_f(g(row, "to be billed")),
                    probability=1.0 if stage == "CloseWon" else (0.5 if stage == "Proposal" else 0.3),
                    notes=_s(g(row, "note")),
                ))
                counts["opportunities"] += 1

    @staticmethod
    def _stage(mms: str) -> str:
        m = (mms or "").lower()
        if "closewon" in m or "won" in m:
            return "CloseWon"
        if "lost" in m:
            return "CloseLost"
        if "proposal" in m or m == "3b":
            return "Proposal"
        if "qualif" in m or m in ("1", "3"):
            return "Qualified"
        return "Lead"

    @staticmethod
    def _fy(title: str) -> str:
        m = re.search(r"FY(\d{2})", title)
        return f"20{m.group(1)}" if m else "2026"
