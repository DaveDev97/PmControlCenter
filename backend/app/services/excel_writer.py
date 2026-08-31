"""Write PM Control Center data back to the Excel workbook.

This module handles updating the Excel file with changes made in the application.
It preserves the existing structure and formatting while updating data.
"""
from __future__ import annotations

import re
from datetime import date
from pathlib import Path

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Contract, Financial, Opportunity, Resource
from app.services.excel_reader import ExcelDataLoader

_CONTRACT_HEADER_RE = re.compile(r"^\s*(\d{6,})\s*-\s*(.+?)\s*$")


def _is_blank(v) -> bool:
    return v is None or v == "" or (isinstance(v, float) and v != v)


class ExcelDataWriter:
    """Writes data back to the security-financials workbook."""

    @staticmethod
    async def save_all(data_folder: Path, session: AsyncSession) -> dict:
        """Save all modified data back to the Excel file."""
        wb_path = ExcelDataLoader.resolve_workbook(data_folder)
        if wb_path is None:
            raise FileNotFoundError(f"No Excel file at {data_folder}")

        # Load workbook (NOT read-only)
        wb = openpyxl.load_workbook(wb_path, data_only=False)

        counts = {"opportunities": 0, "contracts": 0, "financials": 0}

        await _update_opportunities(wb, session, counts)
        # Future: await _update_contracts(wb, session, counts)
        # Future: await _update_financials(wb, session, counts)

        # Save back to file
        wb.save(wb_path)
        return counts


async def _update_opportunities(wb, session: AsyncSession, counts: dict):
    """Update opportunity sheets with current database values."""
    # Get all opportunities from database
    result = await session.execute(select(Opportunity).order_by(Opportunity.fiscal_year, Opportunity.id))
    opportunities = result.scalars().all()

    # Group by fiscal year
    by_fy: dict[str, list[Opportunity]] = {}
    for opp in opportunities:
        fy = opp.fiscal_year or "2026"
        if fy not in by_fy:
            by_fy[fy] = []
        by_fy[fy].append(opp)

    # Update each Opp.FY sheet
    for title in wb.sheetnames:
        if not title.lower().startswith("opp."):
            continue

        ws = wb[title]
        rows = list(ws.iter_rows())

        # Find header row
        hdr_idx = None
        for ri, row in enumerate(rows[:5]):
            if any(cell.value and str(cell.value).strip() == "Contract" for cell in row):
                hdr_idx = ri
                break

        if hdr_idx is None:
            continue

        # Map column names to indices
        cols = {}
        for ci, cell in enumerate(rows[hdr_idx]):
            if cell.value:
                cols[str(cell.value).strip().lower()] = ci

        # Determine FY from sheet title
        fy_match = re.search(r"FY(\d{2})", title)
        fy = f"20{fy_match.group(1)}" if fy_match else "2026"

        # Get opportunities for this FY
        opps_to_write = by_fy.get(fy, [])

        # Update existing rows
        data_start = hdr_idx + 1
        row_idx = data_start

        for opp in opps_to_write:
            if row_idx >= len(rows):
                # Need to add a new row
                ws.append([None] * len(rows[hdr_idx]))
                rows = list(ws.iter_rows())

            row = rows[row_idx]

            # Update cells based on column mapping
            def set_cell(col_name: str, value):
                ci = cols.get(col_name)
                if ci is not None and ci < len(row):
                    row[ci].value = value

            set_cell("project", opp.name)
            set_cell("opp id mms", opp.opp_id_mms)
            set_cell("contract", opp.contract_id)
            set_cell("mms status", _stage_to_mms(opp.stage))
            set_cell("close date", opp.close_date)
            set_cell("close date quarter", opp.quarter)
            set_cell("pds status", opp.pds_status)
            set_cell("stato acn tool", opp.acn_tool_status)
            set_cell("oda id", opp.oda_id)
            set_cell("alphabank contract (ccp)", opp.ccp_number)
            set_cell("bnl contract (ccp)", opp.ccp_number)
            set_cell("ref name", opp.referente)
            set_cell("mmr code", opp.mmr_code)
            set_cell("revenues", opp.estimated_value)
            set_cell("billed", opp.total_invoiced)
            set_cell("to be billed", opp.total_to_invoice)
            set_cell("note", opp.notes)

            row_idx += 1
            counts["opportunities"] += 1


def _stage_to_mms(stage: str) -> str:
    """Convert internal stage to MMS status for Excel."""
    stage_map = {
        "CloseWon": "CloseWon",
        "CloseLost": "CloseLost",
        "Proposal": "3B",
        "Qualified": "1",
        "Lead": "0",
    }
    return stage_map.get(stage, stage)
