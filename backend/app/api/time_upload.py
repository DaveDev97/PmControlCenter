"""Time report upload and ingestion API."""
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook

from app import schemas
from app.core.database import get_session
from app.models import Resource, TimeEntry, Project

router = APIRouter()


@router.post("/upload", response_model=dict)
async def upload_time_report(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
):
    """Upload Excel time report (CHG format) and ingest entries.

    Expected columns: TR, Resource ID, Ore, WBS, Tipologia
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "File must be Excel (.xlsx or .xls)")

    # Read Excel file
    contents = await file.read()
    workbook = load_workbook(BytesIO(contents), data_only=True)
    sheet = workbook.active

    # Parse header row
    headers = [cell.value for cell in sheet[1]]

    # Expected columns
    tr_col = next((i for i, h in enumerate(headers) if h and 'TR' in str(h).upper()), None)
    resource_col = next((i for i, h in enumerate(headers) if h and 'RESOURCE' in str(h).upper()), None)
    ore_col = next((i for i, h in enumerate(headers) if h and 'ORE' in str(h).upper()), None)
    wbs_col = next((i for i, h in enumerate(headers) if h and 'WBS' in str(h).upper()), None)
    tipo_col = next((i for i, h in enumerate(headers) if h and 'TIPO' in str(h).upper()), None)

    if any(col is None for col in [tr_col, resource_col, ore_col, wbs_col, tipo_col]):
        raise HTTPException(400, f"Missing required columns. Found: {headers}")

    # Load all resources by email prefix
    resources_result = await session.execute(select(Resource))
    resources = {r.name.split('@')[0].lower(): r for r in resources_result.scalars().all()}

    # Load all projects by WBS
    projects_result = await session.execute(select(Project))
    projects = {p.wbs: p for p in projects_result.scalars().all()}

    # Parse rows
    entries = []
    errors = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue

        try:
            period = str(row[tr_col]) if row[tr_col] else None
            resource_name = str(row[resource_col]).strip().lower() if row[resource_col] else None
            hours = float(row[ore_col]) if row[ore_col] else 0.0
            wbs = str(row[wbs_col]).strip() if row[wbs_col] else None
            tipo = str(row[tipo_col]).strip() if row[tipo_col] else "Chargeable"

            if not all([period, resource_name, wbs]):
                errors.append(f"Row {row_idx}: Missing required fields")
                continue

            # Find resource
            resource = resources.get(resource_name)
            if not resource:
                errors.append(f"Row {row_idx}: Resource '{resource_name}' not found")
                continue

            # Find project (optional)
            project = projects.get(wbs)

            entries.append({
                'resource_id': resource.id,
                'period': period,
                'hours': hours,
                'wbs': wbs,
                'type': tipo,
                'project_id': project.id if project else None,
            })

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    # Insert entries
    inserted = 0
    for entry_data in entries:
        # Check if entry already exists (avoid duplicates)
        existing = await session.execute(
            select(TimeEntry).where(
                TimeEntry.resource_id == entry_data['resource_id'],
                TimeEntry.period == entry_data['period'],
                TimeEntry.wbs == entry_data['wbs']
            )
        )
        if existing.scalar_one_or_none():
            # Update existing
            existing_entry = existing.scalar_one()
            existing_entry.hours = entry_data['hours']
            existing_entry.type = entry_data['type']
        else:
            # Insert new
            session.add(TimeEntry(**entry_data))
            inserted += 1

    await session.commit()

    return {
        'success': True,
        'inserted': inserted,
        'updated': len(entries) - inserted,
        'errors': errors,
        'total_rows': len(entries) + len(errors)
    }


@router.delete("/entries/{period}", response_model=dict)
async def delete_period_entries(
    period: str,
    session: AsyncSession = Depends(get_session)
):
    """Delete all time entries for a given period (e.g., '2Q-Luglio')."""
    result = await session.execute(
        select(TimeEntry).where(TimeEntry.period == period)
    )
    entries = result.scalars().all()

    for entry in entries:
        await session.delete(entry)

    await session.commit()

    return {
        'success': True,
        'deleted': len(entries),
        'period': period
    }
